import pandas as pd
import re
import os

from config import INPUT_FILE, FILTERED_FILE

def is_gsm_maroc(numero):
    n = re.sub(r"[\s\-\.\(\)\+]", "", str(numero))
    if n.startswith("212"):
        n = "0" + n[3:]
    return bool(re.fullmatch(r"0[67]\d{8}", n))

def normaliser(numero):
    n = re.sub(r"[\s\-\.\(\)\+]", "", str(numero))
    if n.startswith("212"):
        return "+" + n
    elif n.startswith("0"):
        return "+212" + n[1:]
    return n

ext = os.path.splitext(INPUT_FILE)[1].lower()
if ext in [".xlsx", ".xls"]:
    df = pd.read_excel(INPUT_FILE, dtype={"phone": str})
elif ext == ".csv":
    df = pd.read_csv(INPUT_FILE, dtype={"phone": str})
else:
    raise ValueError(f"Format non supporte : {ext}")

print(f"[INFO] {len(df)} lignes chargees")

df["phone"] = df["phone"].astype(str).str.strip()
masque = df["phone"].apply(is_gsm_maroc)

df_valides = df[masque].copy()
df_rejetes = df[~masque].copy()

df_valides["phone"] = df_valides["phone"].apply(normaliser)

print(f"[OK]   {len(df_valides)} numeros valides")
print(f"[REJ]  {len(df_rejetes)} numeros rejetes")

os.makedirs("data", exist_ok=True)
df_valides.to_csv(FILTERED_FILE, index=False)

# ── Excel formaté ──────────────────────────────────────────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

output_xlsx = "data/clients_filtres.xlsx"
with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
    df_valides.to_excel(writer, sheet_name="Valides", index=False)
    df_rejetes.to_excel(writer, sheet_name="Rejetes", index=False)

    wb = writer.book

    def formater_sheet(ws, couleur_header, couleur_row1, couleur_row2):
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 6

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill = PatternFill("solid", start_color=couleur_header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 22

        fills = [PatternFill("solid", start_color=couleur_row1),
                 PatternFill("solid", start_color=couleur_row2)]
        for i, row in enumerate(ws.iter_rows(min_row=2), 1):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.fill = fills[i % 2]
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
            ws.row_dimensions[i + 1].height = 18

    formater_sheet(wb["Valides"], "1F7A4D", "E8F5E9", "FFFFFF")
    formater_sheet(wb["Rejetes"], "C0392B", "FDECEA", "FFFFFF")

print(f"[OK] Excel sauvegarde -> 'data/clients_filtres.xlsx'")
print(f"[OK] CSV sauvegarde   -> '{FILTERED_FILE}'")